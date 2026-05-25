# coding: utf-8
from __future__ import print_function

import glob
import os
import re
import shutil
import subprocess
import sys
from concurrent.futures import ThreadPoolExecutor

import numpy as np
import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


class StudyRunner(object):
    def __init__(self, base_dir=None, max_workers=4):
        self.base_dir = base_dir or os.getcwd()
        self.build_dir = os.path.join(self.base_dir, "build")
        self.results_dir = os.path.join(self.base_dir, "results")
        self.config_file = os.path.join(self.base_dir, "config.cfg")
        self.python = sys.executable
        self.max_workers = max_workers
        self.font = Font(name="Asap", size=8)
        self.font_red = Font(name="Asap", size=8, bold=True, color="FF0000")

    def ensure_workspace(self):
        for folder in (self.build_dir, self.results_dir):
            if not os.path.isdir(folder):
                os.makedirs(folder)

    def call_pssetools(self, command, *args):
        cmd = [self.python, "-m", "pssetools", command]
        cmd.extend(args)
        return subprocess.call(cmd)

    def map_parallel(self, items, worker, *args):
        if not items:
            return []

        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            return list(executor.map(lambda item: worker(item, *args), items))

    @staticmethod
    def scenario_name(case_name):
        return re.sub(r"^.{4}_BESS", "", case_name)

    @staticmethod
    def bess_flag(case_name):
        return "CON" if "CON_BESS" in case_name else "SIN"

    def run_static_case(self, sav, sub, mon, con):
        basename = os.path.basename(sav).replace(".sav", "")
        study_folder = os.path.join(self.results_dir, basename)
        if not os.path.isdir(study_folder):
            os.makedirs(study_folder)

        dfx = os.path.join(self.build_dir, basename + ".dfx")
        acc = os.path.join(self.build_dir, basename + ".acc")
        vrp = os.path.join(self.build_dir, basename + ".vrp")
        frp = os.path.join(self.build_dir, basename + ".frp")
        zipfile = os.path.join(self.build_dir, basename + ".zip")

        self.call_pssetools("dfx", "--sav", sav, "--sub", sub, "--mon", mon, "--con", con, "--dfx", dfx, "--config", self.config_file)
        self.call_pssetools("acc", "--sav", sav, "--dfx", dfx, "--acc", acc, "--zipfile", zipfile, "--config", self.config_file)
        self.call_pssetools("acc-pp", "--sav", sav, "--acc", acc, "--vrp", vrp, "--frp", frp, "--config", self.config_file)
        self.call_pssetools("acc-unzip", "--sav", sav, "--zipfile", zipfile, "--folder", study_folder, "--config", self.config_file)
        return frp, vrp

    def run_static(self, sav_files, sub, mon, con):
        self.ensure_workspace()
        results = self.map_parallel(sav_files, self.run_static_case, sub, mon, con)
        self.write_static_reports(results)

    def write_static_reports(self, results):
        frp_dfs = [pd.read_csv(frp_file, sep="\t") for (frp_file, vrp_file) in results]
        vrp_dfs = [pd.read_csv(vrp_file, sep="\t") for (frp_file, vrp_file) in results]

        pd.concat(frp_dfs).to_excel(os.path.join(self.results_dir, "frp.xlsx"), index=False)
        pd.concat(vrp_dfs).to_excel(os.path.join(self.results_dir, "vfp.xlsx"), index=False)

        vrp = pd.read_excel(os.path.join(self.results_dir, "vfp.xlsx"))
        vrp["ESCENARIO"] = vrp["CASO"].map(self.scenario_name)
        vrp["BESS"] = vrp["CASO"].map(self.bess_flag)
        pt_vrp = vrp.pivot_table("PU", ["ESCENARIO", "ELEMENTO"], ["CONTINGENCIA", "BESS"])
        pt_vrp = pt_vrp[sorted(pt_vrp.columns, key=lambda x: x[0])]

        frp = pd.read_excel(os.path.join(self.results_dir, "frp.xlsx"))
        frp["ESCENARIO"] = frp["CASO"].map(self.scenario_name)
        frp["BESS"] = frp["CASO"].map(self.bess_flag)
        frp["CARGA [%]"] = frp["I [as MVA]"] / frp["LIMITE [MVA]"]
        pt_frp = frp.pivot_table("CARGA [%]", ["ESCENARIO", "ELEMENTO"], ["CONTINGENCIA", "BESS"])
        pt_frp = pt_frp[sorted(pt_frp.columns, key=lambda x: x[0])]

        with pd.ExcelWriter(os.path.join(self.base_dir, "resultados_estatico.xlsx"), engine="openpyxl") as writer:
            for scenario in np.unique([scenario for scenario, _ in pt_frp.index]):
                self.write_static_sheet(writer, "VRP_{}".format(scenario), pt_vrp.loc[scenario], "0.000", self.font)
                self.write_static_sheet(writer, "FRP_{}".format(scenario), pt_frp.loc[scenario], "0%", self.font)

    def write_static_sheet(self, writer, sheet_name, df, number_format, font):
        df.to_excel(writer, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                cell.font = font
                cell.number_format = number_format

        ws.column_dimensions["A"].width = 40 if sheet_name.startswith("VRP") else 50
        for i in range(df.shape[1]):
            ws.column_dimensions[get_column_letter(i + 2)].width = 8 if sheet_name.startswith("VRP") else 6

        start_row = ws.max_row - df.shape[0] + 1
        start_col = ws.max_column - df.shape[1] + 1
        rng = "{}{}:{}{}".format(get_column_letter(start_col), start_row, get_column_letter(ws.max_column), ws.max_row)
        if sheet_name.startswith("VRP"):
            ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0.90"], stopIfTrue=True, font=self.font_red))
            ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["1.10"], stopIfTrue=True, font=self.font_red))
        else:
            ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["1.00"], stopIfTrue=True, font=self.font_red))

    def run_short_circuit_case(self, sav, sub):
        basename = os.path.basename(sav).replace(".sav", "")
        report = os.path.join(self.build_dir, basename + ".scf")
        self.call_pssetools("ascc", "--sav", sav, "--sub", sub, "--report", report, "--config", self.config_file)
        return report

    def run_short_circuit(self, sav_files, sub):
        self.ensure_workspace()
        reports = self.map_parallel(sav_files, self.run_short_circuit_case, sub)
        dfs = [pd.read_csv(report, sep="\t") for report in reports]
        pd.concat(dfs).to_excel(os.path.join(self.results_dir, "scf.xlsx"), index=False)
        self.write_short_circuit_report()

    def write_short_circuit_report(self):
        scf = pd.read_excel(os.path.join(self.results_dir, "scf.xlsx"))
        scf["ESCENARIO"] = scf["CASO"].map(self.scenario_name)
        scf["BESS"] = scf["CASO"].map(self.bess_flag)

        pt_scf = scf.pivot_table(["THREE PHASE FAULT", "LG FAULT"], ["ESCENARIO", "BUS", "NAME", "KV"], ["BESS"])
        pt_scf[("THREE PHASE FAULT", "% DIF")] = pt_scf["THREE PHASE FAULT"]["CON"] / pt_scf["THREE PHASE FAULT"]["SIN"] - 1
        pt_scf[("LG FAULT", "% DIF")] = pt_scf["LG FAULT"]["CON"] / pt_scf["LG FAULT"]["SIN"] - 1
        pt_scf = pt_scf[sorted(pt_scf.columns, key=lambda x: (x[0], x[1]), reverse=True)]

        with pd.ExcelWriter(os.path.join(self.base_dir, "resultados_cortocircuito.xlsx"), engine="openpyxl") as writer:
            pt_scf.to_excel(writer, sheet_name="SCF")
            ws = writer.sheets["SCF"]
            ws.column_dimensions["A"].width = 10
            ws.column_dimensions["B"].width = 10
            ws.column_dimensions["C"].width = 20
            for col in ws.iter_cols():
                for cell in col:
                    cell.font = self.font
                    cell.number_format = "0.00%" if cell.column in (7, 10) else "0"

    def run_scr_case(self, sav, sub):
        basename = os.path.basename(sav).replace(".sav", "")
        foldername = os.path.basename(os.path.split(sav)[0])
        report = os.path.join(self.build_dir, "{}-{}.scf".format(foldername, basename))
        self.call_pssetools("ascc", "--sav", sav, "--sub", sub, "--report", report, "--config", self.config_file)
        return report

    def run_scr(self, sav_files, sub):
        self.ensure_workspace()
        scf = pd.read_excel(os.path.join(self.results_dir, "scf.xlsx"))
        print("barras del sistema:")
        for bus in sorted(scf.BUS.unique()):
            print(bus)

        barra = int(self.ask("ingrese barra a calcular SCR: "))
        potencia = float(self.ask("ingrese potencia (MW): "))
        case_folder = scf[scf.BUS == barra].iloc[np.argmin(scf[scf.BUS == barra]["THREE PHASE FAULT"].values)].CASO
        case_savs = glob.glob(os.path.join(self.results_dir, case_folder, "*.sav"))
        reports = self.map_parallel(case_savs, self.run_scr_case, sub)

        scr_dfs = []
        for report in reports:
            df = pd.read_csv(report, sep="\t")
            df["ESCENARIO"] = os.path.basename(report).split("-")[0]
            scr_dfs.append(df)

        scr_df = pd.concat(scr_dfs)
        scr_df = scr_df[scr_df["BUS"] == barra]
        scr_df["CONTINGENCIA"] = scr_df["CASO"]
        scr_df["CASO"] = case_folder
        scr_df["SCR"] = scr_df["THREE PHASE FAULT"] / potencia
        scr_df = scr_df[["ESCENARIO", "CONTINGENCIA", "SCR"]].set_index(["ESCENARIO", "CONTINGENCIA"]).T

        with pd.ExcelWriter(os.path.join(self.base_dir, "resultados_scr.xlsx"), engine="openpyxl") as writer:
            scr_df.to_excel(writer, sheet_name="SCR")
            ws = writer.sheets["SCR"]
            for col in ws.iter_cols():
                ws.column_dimensions[get_column_letter(col[0].column)].width = 6
                for cell in col:
                    cell.number_format = "0.0"
                    cell.font = self.font

    def clean_results(self):
        print("Limpiando resultados...")
        shutil.rmtree(self.build_dir, ignore_errors=True)
        shutil.rmtree(self.results_dir, ignore_errors=True)
        self.ensure_workspace()

    def ask(self, text):
        try:
            return raw_input(text)
        except NameError:
            return input(text)

    def menu(self):
        self.ensure_workspace()
        print("1. Estatico")
        print("2. Cortocircuito")
        print("3. SCR")
        print("4. Limpiar resultados")
        entrada = str(self.ask("Ingrese la accion a realizar: "))
        sav_files = glob.glob(os.path.join(self.base_dir, "*.sav"))
        sub = "estudio.sub"
        mon = "estudio.mon"
        con = "estudio.con"

        acciones = {
            "1": lambda: self.run_static(sav_files, sub, mon, con),
            "2": lambda: self.run_short_circuit(sav_files, sub),
            "3": lambda: self.run_scr(sav_files, sub),
            "4": self.clean_results,
        }
        if entrada in acciones:
            acciones[entrada]()
        else:
            print("Accion no valida")


if __name__ == "__main__":
    StudyRunner().menu()